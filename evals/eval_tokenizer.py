"""
Tier 1 evaluation of a morphological BPE tokenizer.

Metrics computed:
  1. Morphological boundary precision/recall/F1 against MorphoLex gold standard
  2. Token fertility (avg tokens per word)
  3. Forced morpheme utilization (how often each morpheme token appears)
  4. Compression ratio (bytes per token)
  5. Renyi efficiency (if tokenization-scorer is installed)

Usage:
    # Evaluate a trained tokenizer
    python evals/eval_tokenizer.py --tokenizer_path=tokenizer/output/tokenizer.pkl

    # Compare morpheme vs vanilla tokenizer
    python evals/eval_tokenizer.py \
        --tokenizer_path=tokenizer/output/tokenizer.pkl \
        --baseline_tokenizer=gpt2

    # Evaluate with custom eval text
    python evals/eval_tokenizer.py \
        --tokenizer_path=tokenizer/output/tokenizer.pkl \
        --eval_text=path/to/text.txt
"""

import os
import sys
import re
import json
import pickle
import argparse
from collections import Counter

import tiktoken

# Add parent for tokenizer imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "tokenizer"))

MORPHOLEX_PATH = os.path.join(os.path.dirname(__file__), "MorphoLex-en", "MorphoLEX_en.xlsx")

# ─── MorphoLex parsing ───────────────────────────────────────────────

def parse_morpholex_segmentation(segm_str):
    """
    Parse a MorphoLex segmentation string into morpheme boundary positions.

    Examples:
        '{(algorithm)}>ic>'     -> 'algorithm' + 'ic'  -> boundaries at [9]
        '<un<{(happy)}>ness>'   -> 'un' + 'happy' + 'ness' -> boundaries at [2, 7]
        '{(allow)}>able>>y>'    -> 'allow' + 'able' + 'y' -> boundaries at [5, 9]

    Returns:
        (word, boundaries) where boundaries is a set of character positions
        where morpheme boundaries occur (between char[i-1] and char[i]).
    """
    # Strip the segmentation markers to reconstruct the word and find boundaries
    # Markers: < > { } ( ) and their combinations
    # Strategy: walk through, track when we enter/exit morpheme content

    morphemes = []
    current = []

    i = 0
    s = segm_str
    while i < len(s):
        ch = s[i]
        if ch in '<>{}()':
            # If we have accumulated content, save it as a morpheme
            if current:
                morphemes.append(''.join(current))
                current = []
            i += 1
        else:
            current.append(ch)
            i += 1

    if current:
        morphemes.append(''.join(current))

    # Filter empty morphemes
    morphemes = [m for m in morphemes if m]

    if not morphemes:
        return None, set()

    # Reconstruct word and compute boundary positions
    word = ''.join(morphemes)
    boundaries = set()
    pos = 0
    for m in morphemes[:-1]:  # no boundary after last morpheme
        pos += len(m)
        boundaries.add(pos)

    return word, boundaries


def load_morpholex(path=None):
    """
    Load MorphoLex data, returning a dict of {word: boundaries}.
    Only includes words with 2+ morphemes (so there are boundaries to evaluate).
    """
    if path is None:
        path = MORPHOLEX_PATH

    if not os.path.exists(path):
        print(f"MorphoLex not found at {path}")
        print("Clone it: git clone https://github.com/hugomailhot/MorphoLex-en.git evals/MorphoLex-en")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)

    word_boundaries = {}
    total_words = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in ('Presentation', 'All prefixes', 'All suffixes', 'All roots'):
            continue

        ws = wb[sheet_name]
        # Find the MorphoLexSegm column
        header = None
        segm_col = None
        word_col = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=False)):
            header = [cell.value for cell in row]
            for i, h in enumerate(header):
                if h == 'MorphoLexSegm':
                    segm_col = i
                if h == 'Word':
                    word_col = i

        if segm_col is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[segm_col] is None:
                continue

            total_words += 1
            segm_str = str(row[segm_col])
            parsed_word, boundaries = parse_morpholex_segmentation(segm_str)

            if parsed_word is None or not boundaries:
                skipped += 1
                continue

            # Use the actual Word column value (preserves original casing)
            if word_col is not None and row[word_col] is not None:
                original_word = str(row[word_col]).lower()
            else:
                original_word = parsed_word.lower()

            # Only keep if parsed word matches original (sanity check)
            if original_word == parsed_word.lower():
                word_boundaries[original_word] = boundaries
            else:
                # Try lowercase match
                if original_word.replace("'", "").replace("-", "") == parsed_word.lower().replace("'", "").replace("-", ""):
                    word_boundaries[original_word] = boundaries
                else:
                    skipped += 1

    wb.close()
    return word_boundaries


# ─── Evaluation metrics ──────────────────────────────────────────────

def get_tokenizer_boundaries(enc, word):
    """
    Tokenize a word and return the character positions where token boundaries fall.

    For a word tokenized as [tok1, tok2, tok3], boundaries are at positions
    [len(tok1_text), len(tok1_text)+len(tok2_text)].
    """
    ids = enc.encode(word)
    boundaries = set()
    pos = 0
    for token_id in ids[:-1]:  # no boundary after last token
        token_bytes = enc.decode([token_id]).encode("utf-8")
        pos += len(token_bytes)
        # Convert byte position to character position
        # (for ASCII words these are the same)
        boundaries.add(pos)

    return boundaries


def compute_boundary_metrics(enc, word_boundaries, max_words=None):
    """
    Compute precision/recall/F1 of tokenizer boundaries vs gold morpheme boundaries.

    For each word:
      - TP: boundary exists in both tokenizer and gold
      - FP: boundary in tokenizer but not in gold
      - FN: boundary in gold but not in tokenizer
    """
    tp = fp = fn = 0
    evaluated = 0

    for word, gold_boundaries in word_boundaries.items():
        if max_words and evaluated >= max_words:
            break

        try:
            tok_boundaries = get_tokenizer_boundaries(enc, word)
        except Exception:
            continue

        evaluated += 1
        for b in tok_boundaries:
            if b in gold_boundaries:
                tp += 1
            else:
                fp += 1
        for b in gold_boundaries:
            if b not in tok_boundaries:
                fn += 1

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    return {
        "precision": precision,
        "recall": recall,
        "f1": f1,
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "words_evaluated": evaluated,
    }


def compute_fertility(enc, words):
    """Compute average tokens per word."""
    total_tokens = 0
    total_words = 0
    for word in words:
        try:
            ids = enc.encode(word)
            total_tokens += len(ids)
            total_words += 1
        except Exception:
            continue

    return {
        "avg_tokens_per_word": total_tokens / total_words if total_words > 0 else 0,
        "total_tokens": total_tokens,
        "total_words": total_words,
    }


def compute_compression_ratio(enc, text):
    """Compute bytes per token on a text sample."""
    ids = enc.encode(text)
    text_bytes = len(text.encode("utf-8"))
    return {
        "bytes_per_token": text_bytes / len(ids) if ids else 0,
        "total_bytes": text_bytes,
        "total_tokens": len(ids),
    }


def compute_morpheme_utilization(enc, text, morphemes, meta=None):
    """
    Count how often each forced morpheme token appears when encoding text.
    Returns utilization stats per morpheme.
    """
    ids = enc.encode(text)
    id_counts = Counter(ids)

    # Build morpheme -> expected token ID mapping
    results = []
    for morpheme in morphemes:
        # Encode the morpheme alone to find its token ID
        morph_ids = enc.encode(morpheme)
        is_single_token = len(morph_ids) == 1
        morph_token_id = morph_ids[-1] if morph_ids else None
        count = id_counts.get(morph_token_id, 0) if morph_token_id else 0

        results.append({
            "morpheme": morpheme,
            "token_id": morph_token_id,
            "is_single_token": is_single_token,
            "count_in_corpus": count,
        })

    # Sort by count descending
    results.sort(key=lambda x: x["count_in_corpus"], reverse=True)

    total_used = sum(1 for r in results if r["count_in_corpus"] > 0)
    total_morphemes = len(morphemes)

    return {
        "morphemes": results,
        "total_morphemes": total_morphemes,
        "morphemes_used": total_used,
        "morphemes_unused": total_morphemes - total_used,
        "utilization_rate": total_used / total_morphemes if total_morphemes > 0 else 0,
    }


# ─── Main ────────────────────────────────────────────────────────────

def load_tokenizer(path):
    """Load a tiktoken encoding from a pickle file."""
    with open(path, "rb") as f:
        return pickle.load(f)


DEFAULT_EVAL_TEXT = """
The unhappiness of the disconnected international community was unreasonable.
They were rethinking their preexisting assumptions about interconnected systems.
The disorganization was overwhelming and the misinformation spread quickly.
Scientists were rediscovering fundamental truths about the natural world.
The government was reorganizing its departments for greater effectiveness.
Teachers found that rewriting the curriculum improved student engagement.
The remarkable transformation of the industry was unprecedented in history.
Understanding the relationship between education and innovation requires careful analysis.
The international community expressed disagreement about environmental regulations.
Researchers were investigating the interconnection of biological systems.
The unreliable predictions were inconsistent with the experimental observations.
Pharmaceutical companies were developing new treatments for autoimmune disorders.
The decentralization of power led to greater accountability and transparency.
Environmentalists were campaigning against the deforestation of tropical rainforests.
The incomprehensible complexity of the neural network amazed the researchers.
Multinational corporations were restructuring their operations internationally.
The irresponsible behavior of the politicians was widely criticized.
Archaeologists made an extraordinary discovery about prehistoric civilizations.
The counterproductive policies were eventually reversed by the new administration.
""" * 20


def main():
    parser = argparse.ArgumentParser(description="Evaluate a morphological BPE tokenizer")
    parser.add_argument("--tokenizer_path", type=str, required=True,
                        help="Path to trained tokenizer pickle")
    parser.add_argument("--baseline_tokenizer", type=str, default=None,
                        help="Baseline tiktoken encoding name (e.g. 'gpt2') for comparison")
    parser.add_argument("--eval_text", type=str, default=None,
                        help="Path to evaluation text file (for fertility/compression)")
    parser.add_argument("--max_words", type=int, default=None,
                        help="Max MorphoLex words to evaluate (for speed)")
    parser.add_argument("--output", type=str, default=None,
                        help="Save results to JSON file")
    args = parser.parse_args()

    results = {}

    # Load tokenizer
    print(f"Loading tokenizer from {args.tokenizer_path}")
    enc = load_tokenizer(args.tokenizer_path)

    # Load metadata if available
    meta_path = os.path.join(os.path.dirname(args.tokenizer_path), "meta.json")
    meta = None
    morphemes = None
    if os.path.exists(meta_path):
        with open(meta_path) as f:
            meta = json.load(f)
        morphemes = meta.get("forced_morphemes")
        print(f"Vocab size: {meta.get('vocab_size')}, "
              f"Forced merges: {meta.get('forced_merge_count')}")

    # Load baseline if requested
    baseline_enc = None
    if args.baseline_tokenizer:
        print(f"Loading baseline: {args.baseline_tokenizer}")
        baseline_enc = tiktoken.get_encoding(args.baseline_tokenizer)

    # Load eval text
    if args.eval_text:
        with open(args.eval_text, "r") as f:
            eval_text = f.read()
    else:
        eval_text = DEFAULT_EVAL_TEXT

    # ── Metric 1: Morphological boundary P/R/F1 ──
    print("\n" + "=" * 60)
    print("METRIC 1: Morphological Boundary Alignment (MorphoLex)")
    print("=" * 60)
    if os.path.exists(MORPHOLEX_PATH):
        print("Loading MorphoLex...")
        word_boundaries = load_morpholex()
        print(f"Loaded {len(word_boundaries)} words with morpheme boundaries")

        print("\nMorpheme tokenizer:")
        metrics = compute_boundary_metrics(enc, word_boundaries, args.max_words)
        results["morpheme_boundary"] = metrics
        print(f"  Precision: {metrics['precision']:.4f}")
        print(f"  Recall:    {metrics['recall']:.4f}")
        print(f"  F1:        {metrics['f1']:.4f}")
        print(f"  (TP={metrics['tp']}, FP={metrics['fp']}, FN={metrics['fn']}, "
              f"words={metrics['words_evaluated']})")

        if baseline_enc:
            print(f"\nBaseline ({args.baseline_tokenizer}):")
            baseline_metrics = compute_boundary_metrics(baseline_enc, word_boundaries, args.max_words)
            results["baseline_boundary"] = baseline_metrics
            print(f"  Precision: {baseline_metrics['precision']:.4f}")
            print(f"  Recall:    {baseline_metrics['recall']:.4f}")
            print(f"  F1:        {baseline_metrics['f1']:.4f}")

            delta_f1 = metrics['f1'] - baseline_metrics['f1']
            print(f"\n  Delta F1: {delta_f1:+.4f}")
    else:
        print("MorphoLex not found — skipping. Clone it:")
        print("  git clone https://github.com/hugomailhot/MorphoLex-en.git evals/MorphoLex-en")

    # ── Metric 2: Token Fertility ──
    print("\n" + "=" * 60)
    print("METRIC 2: Token Fertility")
    print("=" * 60)
    # Use MorphoLex words if available, else extract words from eval text
    if 'word_boundaries' in dir() and word_boundaries:
        fertility_words = list(word_boundaries.keys())[:10000]
    else:
        fertility_words = list(set(re.findall(r'\b[a-zA-Z]+\b', eval_text)))

    print(f"\nMorpheme tokenizer ({len(fertility_words)} words):")
    fertility = compute_fertility(enc, fertility_words)
    results["fertility"] = fertility
    print(f"  Avg tokens/word: {fertility['avg_tokens_per_word']:.3f}")

    if baseline_enc:
        print(f"\nBaseline ({args.baseline_tokenizer}):")
        baseline_fertility = compute_fertility(baseline_enc, fertility_words)
        results["baseline_fertility"] = baseline_fertility
        print(f"  Avg tokens/word: {baseline_fertility['avg_tokens_per_word']:.3f}")

    # ── Metric 3: Compression Ratio ──
    print("\n" + "=" * 60)
    print("METRIC 3: Compression Ratio")
    print("=" * 60)
    print(f"\nMorpheme tokenizer:")
    compression = compute_compression_ratio(enc, eval_text)
    results["compression"] = compression
    print(f"  Bytes/token: {compression['bytes_per_token']:.2f}")
    print(f"  Total: {compression['total_bytes']:,} bytes -> {compression['total_tokens']:,} tokens")

    if baseline_enc:
        print(f"\nBaseline ({args.baseline_tokenizer}):")
        baseline_compression = compute_compression_ratio(baseline_enc, eval_text)
        results["baseline_compression"] = baseline_compression
        print(f"  Bytes/token: {baseline_compression['bytes_per_token']:.2f}")

    # ── Metric 4: Forced Morpheme Utilization ──
    if morphemes:
        print("\n" + "=" * 60)
        print("METRIC 4: Forced Morpheme Utilization")
        print("=" * 60)
        util = compute_morpheme_utilization(enc, eval_text, morphemes, meta)
        results["morpheme_utilization"] = {
            "total": util["total_morphemes"],
            "used": util["morphemes_used"],
            "unused": util["morphemes_unused"],
            "rate": util["utilization_rate"],
        }
        print(f"  Used: {util['morphemes_used']}/{util['total_morphemes']} "
              f"({util['utilization_rate']:.1%})")

        print(f"\n  Top 15 most used:")
        for m in util["morphemes"][:15]:
            status = "single" if m["is_single_token"] else "MULTI"
            print(f"    {m['morpheme']:>10s}  id={m['token_id']:>6}  "
                  f"count={m['count_in_corpus']:>6}  [{status}]")

        unused = [m for m in util["morphemes"] if m["count_in_corpus"] == 0]
        if unused:
            print(f"\n  Unused morphemes ({len(unused)}):")
            for m in unused:
                print(f"    {m['morpheme']}")

    # ── Metric 5: Renyi Efficiency ──
    print("\n" + "=" * 60)
    print("METRIC 5: Renyi Efficiency")
    print("=" * 60)
    try:
        from tokenization_scorer import score
        ids = enc.encode(eval_text)
        renyi = score(ids, alpha=2.5)
        results["renyi_efficiency"] = renyi
        print(f"  Renyi efficiency (alpha=2.5): {renyi:.4f}")

        if baseline_enc:
            baseline_ids = baseline_enc.encode(eval_text)
            baseline_renyi = score(baseline_ids, alpha=2.5)
            results["baseline_renyi"] = baseline_renyi
            print(f"  Baseline Renyi: {baseline_renyi:.4f}")
    except ImportError:
        print("  tokenization-scorer not installed. Install with:")
        print("  pip install tokenization-scorer")
        results["renyi_efficiency"] = None

    # ── Summary ──
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    if "morpheme_boundary" in results:
        print(f"  Boundary F1:     {results['morpheme_boundary']['f1']:.4f}", end="")
        if "baseline_boundary" in results:
            delta = results['morpheme_boundary']['f1'] - results['baseline_boundary']['f1']
            print(f"  (baseline: {results['baseline_boundary']['f1']:.4f}, delta: {delta:+.4f})")
        else:
            print()
    print(f"  Fertility:       {results.get('fertility', {}).get('avg_tokens_per_word', 'N/A'):.3f} tokens/word", end="")
    if "baseline_fertility" in results:
        print(f"  (baseline: {results['baseline_fertility']['avg_tokens_per_word']:.3f})")
    else:
        print()
    print(f"  Compression:     {results.get('compression', {}).get('bytes_per_token', 'N/A'):.2f} bytes/token", end="")
    if "baseline_compression" in results:
        print(f"  (baseline: {results['baseline_compression']['bytes_per_token']:.2f})")
    else:
        print()
    if morphemes and "morpheme_utilization" in results:
        print(f"  Morph. utilized: {results['morpheme_utilization']['rate']:.1%} "
              f"({results['morpheme_utilization']['used']}/{results['morpheme_utilization']['total']})")

    # Save results
    if args.output:
        with open(args.output, "w") as f:
            json.dump(results, f, indent=2)
        print(f"\nResults saved to {args.output}")


if __name__ == "__main__":
    main()
